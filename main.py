import requests
import pandas as pd
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import smtplib
from email.message import EmailMessage
import os

# =========================================
# CONFIG
# =========================================

BASE_API = "https://services-beta.vipcommerce.com.br"

LOJA_URL = "https://www.spanionline.com.br"

BUSCA = "a"

LIMITE_ITENS = 999999

OUTPUT = "SPANI.xlsx"

EMAIL_USER = os.getenv("EMAIL_USER")

EMAIL_PASS = os.getenv("EMAIL_PASS")

EMAIL_TO = "luciano.nunes@roldao.com.br, evedson.felix@roldao.com.br, leandro.moreira@roldao.com.br, ali.sati@roldao.com.br, gabriela.novais@roldao.com.br, gilmar.lana@roldao.com.br, gustavo.henrique@roldao.com.br, marcos.teixeira@roldao.com.br, pricing@roldao.com.br, comercial@roldao.com.br, gerente.regional@roldao.com.br, guilherme.roldao@roldao.com.br, victoria.roldao@roldao.com.br"

# =========================================
# SETORES
# =========================================

SETORES = {

    15: "PET",
    16: "PET",
    17: "PET",
    19: "UTILIDADES",
    20: "DESCARTAVEIS",
    21: "BAZAR",
    22: "BAZAR",
    24: "UTILIDADES",
    25: "LIMPEZA",
    26: "AUTOMOTIVO",
    27: "PAPELARIA",
    28: "DESCARTAVEIS",
    29: "PILHAS",
    30: "LIMPEZA",
    32: "UTILIDADES",
    33: "ELETRICOS",
    34: "REFRIGERANTE",
    35: "CHA",
    38: "SUCOS",
    40: "APERITIVOS",
    41: "BEBIDAS",
    43: "VINHO",
    44: "BEBIDAS",
    45: "CERVEJA",
    46: "BEBIDAS",
    47: "ENERGETICOS",
    48: "BEBIDAS",
    49: "ISOTONICOS",
    50: "BEBIDAS",
    51: "BISCOITOS",
    52: "BISCOITOS",
    53: "DOCES",
    54: "SALGADINHOS",
    55: "BOMBONS",
    57: "AMENDOINS",
    58: "BISCOITOS",
    59: "CHOCOLATES",
    61: "TORRADAS",
    62: "BISCOITOS",
    63: "BISCOITOS",
    65: "DOCES",
    66: "BISCOITOS",
    67: "BISCOITOS",
    68: "UTILIDADES",
    69: "CONGELADOS",
    70: "FRIOS",
    71: "ACOUGUE",
    72: "FRIOS",
    73: "PEIXARIA",
    74: "ACOUGUE",
    75: "ACOUGUE",
    76: "ACOUGUE",
    77: "FRIOS",
    80: "MERCEARIA",
    81: "MERCEARIA",
    82: "MERCEARIA",
    83: "MERCEARIA",
    84: "MERCEARIA",
    86: "CONGELADOS",
    87: "MERCEARIA",
    88: "MERCEARIA",
    92: "MERCEARIA",
    93: "MERCEARIA",
    95: "CONGELADOS",
    96: "CONGELADOS",
    97: "CONGELADOS",
    98: "CONGELADOS",
    99: "CONGELADOS",
    100: "CONGELADOS",
    101: "CONGELADOS",
    102: "CONGELADOS",
    103: "CONGELADOS",
    104: "LATICINIOS",
    105: "LATICINIOS",
    107: "LATICINIOS",
    108: "LATICINIOS",
    109: "LEITE",
    110: "MASSAS",
    111: "LATICINIOS",
    112: "FRIOS",
    113: "LATICINIOS",
    114: "FRIOS",
    115: "FRIOS",
    116: "LATICINIOS",
    117: "HORTIFRUTI",
    118: "HORTIFRUTI",
    119: "HORTIFRUTI",
    120: "HORTIFRUTI",
    121: "HORTIFRUTI",
    123: "LIMPEZA",
    124: "LIMPEZA",
    125: "LIMPEZA",
    126: "LIMPEZA",
    127: "LIMPEZA",
    129: "LIMPEZA",
    130: "LIMPEZA",
    131: "VESTUARIO",
    132: "LIMPEZA",
    133: "LIMPEZA",
    137: "LIMPEZA",
    138: "MERCEARIA",
    139: "CEREAIS",
    140: "CAFE",
    141: "MERCEARIA",
    142: "SUPLEMENTOS",
    143: "MERCEARIA",
    144: "MERCEARIA",
    145: "ACHOCOLATADOS",
    146: "LEITE",
    148: "MERCEARIA",
    150: "CEREAIS",
    151: "MERCEARIA",
    152: "MERCEARIA",
    153: "CHA",
    154: "ACUCAR",
    155: "MERCEARIA",
    156: "INFANTIL",
    158: "MACARRAO",
    159: "MERCEARIA",
    160: "TEMPEROS",
    161: "OLEOS",
    162: "MACARRAO",
    163: "OLEOS",
    164: "MERCEARIA",
    165: "MERCEARIA",
    166: "MERCEARIA",
    167: "MERCEARIA",
    168: "MERCEARIA",
    170: "TEMPEROS",
    171: "MACARRAO",
    172: "MACARRAO",
    173: "MERCEARIA",
    174: "MERCEARIA",
    175: "MERCEARIA",
    176: "MERCEARIA",
    177: "MERCEARIA",
    178: "MACARRAO",
    179: "MACARRAO",
    180: "PADARIA",
    181: "PADARIA",
    182: "PADARIA",
    183: "BAZAR",
    184: "HIGIENE",
    185: "HIGIENE",
    186: "HIGIENE",
    188: "HIGIENE",
    191: "HIGIENE",
    192: "INFANTIL",
    193: "HIGIENE",
    194: "HIGIENE",
    195: "INFANTIL",
    196: "INFANTIL",
    197: "HIGIENE",
    198: "HIGIENE",
    199: "HIGIENE",
    200: "HIGIENE",
    201: "PAPEL HIGIENICO",
    202: "HIGIENE",
    205: "HIGIENE",
    206: "HIGIENE",
    208: "LIMPEZA",
    210: "FRALDAS",
    211: "CONGELADOS",
    215: "CONGELADOS",
    218: "BISCOITOS",
    219: "BISCOITOS",
    220: "CERVEJA",
    221: "HIGIENE",
    222: "LIMPEZA",
    223: "LIMPEZA",
    224: "MERCEARIA",
    225: "HIGIENE",
    226: "ESPUMANTES",
    227: "MERCEARIA",
    228: "MERCEARIA",
    229: "INFANTIL",
    230: "MERCEARIA",
    231: "MERCEARIA",
    232: "LIMPEZA",
    233: "LATICINIOS",
    234: "LATICINIOS",
    235: "DESCARTAVEIS",
    236: "MERCEARIA",
    237: "BEBIDAS",
    238: "BEBIDAS",
    239: "HIGIENE",
    240: "LIMPEZA",
    242: "CONGELADOS",
    243: "FRALDAS",
    244: "MERCEARIA",
    245: "LIMPEZA",
    246: "LATICINIOS",
    248: "MERCEARIA"
}

# =========================================
# FUNCAO PRECO
# =========================================

def formatar_preco(valor):

    if valor is None or valor == "":

        return ""

    try:

        return (
            f"{float(str(valor).replace(',', '.')):.2f}"
            .replace(".", ",")
        )

    except:

        return str(valor)

# =========================================
# PLAYWRIGHT
# =========================================

print("ABRINDO SPANI...")

with sync_playwright() as p:

    browser = p.chromium.launch(
        headless=True
    )

    context = browser.new_context()

    page = context.new_page()

    page.goto(
        LOJA_URL,
        timeout=120000
    )

    page.wait_for_timeout(10000)

    print("SELECIONANDO MAUA 1")

    page.locator(
        ".vip-endereco-wrapper"
    ).click()

    page.wait_for_timeout(3000)

    page.locator(
        "text=Spani Mauá 1"
    ).click()

    page.wait_for_timeout(10000)

    cookies_play = context.cookies()

    session_id = ""

    vip_token = ""

    for c in cookies_play:

        if c["name"] == "sessao-id":

            session_id = c["value"]

        if c["name"] == "vip-token":

            vip_token = c["value"]

    browser.close()

# =========================================
# HEADERS
# =========================================

headers = {

    "accept": "application/json",

    "origin": "https://www.spanionline.com.br",

    "referer": "https://www.spanionline.com.br/",

    "Authorization": f"Bearer {vip_token}",

    "organizationid": "67",

    "filialid": "1",

    "centrodistribuicaoid": "36",

    "user-agent": "Mozilla/5.0"
}

cookies = {

    "sessao-id": session_id,

    "vip-token": vip_token
}

# =========================================
# BUSCA
# =========================================

pagina = 1

todos = []

while True:

    if len(todos) >= LIMITE_ITENS:

        break

    url = (

        f"{BASE_API}"

        f"/api-admin/v1/org/67"

        f"/filial/1"

        f"/centro_distribuicao/36"

        f"/loja/buscas/produtos/termo/{BUSCA}"

        f"?page={pagina}"

        f"&session={session_id}"
    )

    print(f"PAGINA {pagina}")

    r = requests.get(

        url,

        headers=headers,

        cookies=cookies,

        timeout=120
    )

    print("STATUS:", r.status_code)

    if r.status_code != 200:

        print(r.text)

        break

    data = r.json()

    produtos = data.get(
        "data",
        {}
    ).get(
        "produtos",
        []
    )

    if len(produtos) == 0:

        break

    for p in produtos:

        try:

            secao_id = p.get(
                "secao_id",
                0
            )

            setor = SETORES.get(

                secao_id,

                f"SETOR {secao_id}"
            )

            produto = (

                p.get(
                    "descricao",
                    ""
                )

                .strip()

                .upper()
            )

            ean = p.get(
                "codigo_barras",
                ""
            )

            varejo_valor = p.get(
                "preco",
                ""
            )

            varejo = formatar_preco(
                varejo_valor
            )

            atacado = ""

            qtd_atacado = ""

            oferta = p.get("oferta")

            if oferta:

                preco_oferta = oferta.get(
                    "preco_oferta"
                )

                quantidade_minima = oferta.get(
                    "quantidade_minima"
                )

                try:

                    varejo_float = float(varejo_valor)

                    atacado_float = float(preco_oferta)

                    if atacado_float < varejo_float:

                        atacado = formatar_preco(
                            preco_oferta
                        )

                        qtd_atacado = quantidade_minima

                except:

                    pass

            slug = p.get(
                "link",
                ""
            )

            produto_id = p.get(
                "produto_id",
                ""
            )

            link = (

                "https://www.spanionline.com.br/produto/"

                f"{produto_id}/"

                f"{slug}"
            )

            todos.append({

                "SETOR": setor,

                "PRODUTO": produto,

                "VAREJO": varejo,

                "ATACADO": atacado,

                "QTD ATACADO": qtd_atacado,

                "EAN": ean,

                "LINK": link
            })

        except Exception as e:

            print(
                "ERRO:",
                e
            )

    pagina += 1

# =========================================
# VALIDAR DADOS
# =========================================

if len(todos) == 0:

    raise Exception(
        "SEM DADOS COLETADOS"
    )

# =========================================
# DATAFRAME
# =========================================

df = pd.DataFrame(todos)

df = df.sort_values(
    by=["SETOR", "PRODUTO"]
).reset_index(drop=True)

print(df.head())

# =========================================
# EXCEL
# =========================================

df.to_excel(
    OUTPUT,
    index=False
)

wb = load_workbook(
    OUTPUT
)

ws = wb.active

ws.freeze_panes = "A2"

fill = PatternFill(

    start_color="16365C",

    end_color="16365C",

    fill_type="solid"
)

font = Font(

    color="FFFFFF",

    bold=True
)

for cell in ws[1]:

    cell.fill = fill

    cell.font = font

for row in range(2, ws.max_row + 1):

    cell = ws[f"G{row}"]

    url = cell.value

    cell.value = "ABRIR"

    cell.hyperlink = url

    cell.style = "Hyperlink"

larguras = {

    1: 28,
    2: 70,
    3: 12,
    4: 12,
    5: 15,
    6: 20,
    7: 12
}

for col, largura in larguras.items():

    ws.column_dimensions[
        get_column_letter(col)
    ].width = largura

tab = Table(

    displayName="TabelaSpani",

    ref=f"A1:G{ws.max_row}"
)

style = TableStyleInfo(

    name="TableStyleMedium2",

    showRowStripes=False,

    showColumnStripes=False
)

tab.tableStyleInfo = style

ws.add_table(tab)

wb.save(OUTPUT)

print("EXCEL FINALIZADO")

# =========================================
# EMAIL
# =========================================

try:

    msg = EmailMessage()

    msg["Subject"] = "SPANI - RELATORIO PRECOS"

    msg["From"] = EMAIL_USER

    msg["To"] = EMAIL_TO

    msg.set_content("""

Boa Tarde,

Segue em anexo o relatório atualizado de preços coletados no site do Spani Atacadista.

A coleta foi realizada considerando a loja Spani Mauá 1, localizada em Mauá/SP.

Arquivo gerado automaticamente pelo robô de monitoramento de preços.

Att,
Bruno

""")

    with open(
        OUTPUT,
        "rb"
    ) as f:

        msg.add_attachment(

            f.read(),

            maintype="application",

            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",

            filename=OUTPUT
        )

    with smtplib.SMTP(
        "smtp.gmail.com",
        587
    ) as smtp:

        smtp.starttls()

        smtp.login(
            EMAIL_USER,
            EMAIL_PASS
        )

        smtp.send_message(msg)

    print("EMAIL ENVIADO")

except Exception as e:

    print(
        "ERRO EMAIL:",
        e
    )

print("FINALIZADO")
